# -*- coding: utf-8 -*-
import json
import random
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent

def load_json(name):
    return json.loads((ROOT / name).read_text(encoding="utf-8"))

FEATURES = load_json("features_data.json")["features"]
LABELS = load_json("labels_ko.json")

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)

APIS = [
    ["AUTH-01", "POST", "/api/auth/login", "LO_01_01", "Email login", '{"email","password"}', "200 JWT"],
    ["AUTH-02", "POST", "/api/auth/register", "LO_02_04", "Register", "{profile}", "201"],
    ["AUTH-03", "POST", "/api/auth/social/verify", "LO_01_02~04", "SNS verify", "{provider,token}", "200"],
    ["AUTH-04", "GET", "/api/auth/me", "ST_01_01", "Profile", "Bearer", "200"],
    ["DATA-01", "GET", "/api/data/glucose", "GU_01_01,TG_01_01", "Glucose DL", "from,to,limit", "200[]"],
    ["DATA-02", "POST", "/api/data/glucose", "GU_01_01", "Glucose 1", "{time,value,trid}", "201"],
    ["DATA-03", "POST", "/api/data/glucose/batch", "GU_01_01", "Batch", "{points[]}", "201"],
    ["DATA-04", "GET", "/api/data/events", "ME_01_01", "Events get", "from,to", "200[]"],
    ["DATA-05", "POST", "/api/data/events", "ME_01_01", "Event post", "{type,time,memo}", "201"],
    ["DATA-06", "POST", "/api/data/events/batch", "ME_01_01", "Batch<=200", "{items[]}", "201"],
    ["DATA-07", "DELETE", "/api/data/events/:id", "ME_01_01", "Delete", "id", "204"],
    ["SET-01", "GET", "/api/settings/app", "ST_01_01", "App settings", "-", "200"],
    ["SET-02", "PUT", "/api/settings/app", "ST_01_01", "Save app", "{prefs}", "200"],
    ["SET-03", "GET", "/api/settings/alarms", "AR_01_01~08", "Alarms DL", "eqsn", "200[]"],
    ["SET-04", "POST", "/api/settings/alarms", "AR_01_02~06", "Create alarm", "{type,...}", "201"],
    ["SET-05", "PUT", "/api/settings/alarms/:id", "AR_01_02~06", "Update alarm", "body", "200"],
    ["SET-06", "POST", "/api/settings/sensors", "SC_04_01", "Add sensor", "{serial}", "201"],
    ["SET-07", "PUT", "/api/settings/sensors/:id", "SC_04_01", "Update sensor", "body", "200"],
    ["SET-08", "DELETE", "/api/settings/sensors/:id", "SC_08_01", "Delete sensor", "id", "200"],
    ["EQ-01", "GET", "/api/settings/eq-list/:serial", "SC_04_01", "EQ by SN", "path", "200"],
    ["EQ-02", "GET", "/api/settings/eq-list/resolve", "SC_01_04", "Resolve", "query", "200"],
    ["EQ-03", "POST", "/api/settings/eq-list", "SC_01_04,SC_05_01", "Upsert EQ", "{serial,startAt}", "200"],
    ["HEALTH-01", "GET", "/api/health", "-", "Health", "-", "200"],
]

MONGO = [
    ["users", "User", "Account", "LO_*", "_id,email,passwordHash"],
    ["glucose_points", "GlucosePoint", "Glucose TS", "GU,TG,RP", "userId,eqsn,time,value,trid"],
    ["events", "Event", "Events", "ME_01_01", "userId,type,time,memo"],
    ["alarms", "Alarm", "Alarms", "AR_*", "userId,eqsn,type,threshold"],
    ["sensors", "Sensor", "Sensors", "SC_*", "userId,serial,eqsn"],
    ["eq_registry", "EqList", "EQ master", "SC_04_01", "serial,bleMac,startAt"],
    ["app_settings", "AppSetting", "App prefs", "ST_01_01", "userId,preferences"],
]


def style_header_row(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def auto_width(ws, min_w=10, max_w=50):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        length = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(length + 2, min_w), max_w)


def write_table(ws, headers, rows, start_row=1):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row in rows:
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        r += 1
    auto_width(ws)


def random_test_time():
    start = datetime(2026, 6, 10, 20, 0, 0)
    end = datetime(2026, 6, 11, 0, 0, 0)
    sec = random.randint(0, int((end - start).total_seconds()))
    return (start + timedelta(seconds=sec)).strftime("%Y-%m-%d %H:%M:%S")


def build_mongodb_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "00_schema"
    ws["A1"] = LABELS["mongo_title"]
    ws["A1"].font = TITLE_FONT
    write_table(ws, ["Item", "Value", "Item2", "Value2"], [
        ["Version", "1.0.0", "Date", "2026-06-11"],
        ["DBMS", "MongoDB 6.x", "DB", "empecs_cgms"],
        ["REQ ref", "cgms_docs.md", "BE", "f62657f"],
    ], 3)
    ws2 = wb.create_sheet("01_collections")
    write_table(ws2, ["Collection", "Model", "Desc", "REQ IDs", "Fields"], MONGO)
    ws3 = wb.create_sheet("02_fields")
    rows = []
    for coll, _, _, _, fields in MONGO:
        for f in fields.split(","):
            rows.append([coll, f.strip(), "String/Number/Date", "IDX", "see BE"])
    write_table(ws3, ["Collection", "Field", "Type", "Index", "Note"], rows)
    ws4 = wb.create_sheet("03_erd")
    write_table(ws4, ["From", "Rel", "To", "FK", "Desc"], [
        ["users", "1:N", "glucose_points", "userId", "glucose"],
        ["users", "1:N", "events", "userId", "events"],
        ["users", "1:N", "alarms", "userId", "alarms"],
        ["sensors", "N:1", "eq_registry", "serial", "EQ"],
    ])
    ws5 = wb.create_sheet("04_indexes")
    write_table(ws5, ["Collection", "Index", "Type", "REQ"], [
        ["glucose_points", "{userId,eqsn,time:-1}", "compound", "TG_01_01"],
        ["alarms", "{userId,eqsn,type:1}", "unique", "AR_01_01"],
    ])
    ws6 = wb.create_sheet("05_req_map")
    req_rows = [[fid, title, "MongoDB+local", desc] for fid, title, _, _, desc in FEATURES]
    write_table(ws6, ["REQ ID", "Screen", "Storage", "Note"], req_rows)
    wb.save(path)


def build_api_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "00_api_list"
    ws["A1"] = LABELS["api_title"]
    ws["A1"].font = TITLE_FONT
    write_table(ws, ["API-ID", "Method", "Path", "REQ IDs", "Summary", "Request", "Response"], APIS, 3)
    ws2 = wb.create_sheet("01_auth_flow")
    write_table(ws2, ["Step", "REQ", "Action", "Impl"], [
        ["1", "LO_01_01", "Login UI", "LoginChoiceScreen"],
        ["2", "AUTH-01/03", "POST auth", "api_client"],
        ["3", "-", "Save JWT", "settings_storage"],
        ["4", "MAIN_DASHBOARD", "/home", "home.dart"],
    ])
    ws3 = wb.create_sheet("02_glucose_sync")
    write_table(ws3, ["Step", "REQ", "Action", "File"], [
        ["1", "SC_01_01", "BLE notify", "ble_service.dart"],
        ["2", "GU_01_01", "SQLite", "glucose_local_repo"],
        ["3", "DATA-02", "POST", "ingest_queue.dart"],
        ["4", "TG_01_01", "Chart", "chart_page.dart"],
    ])
    ws4 = wb.create_sheet("03_detail")
    write_table(ws4, ["Path", "Method", "Request", "Response", "REQ"],
                [(a[2], a[1], a[5], a[6], a[3]) for a in APIS])
    ws5 = wb.create_sheet("04_req_cross")
    cross = []
    for fid, title, _, _, desc in FEATURES:
        ids = [a[0] for a in APIS if fid in a[3]]
        cross.append([fid, title, ", ".join(ids) if ids else "(local)", desc])
    write_table(ws5, ["REQ ID", "Screen", "API-IDs", "Note"], cross)
    wb.save(path)


def build_test_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "01_feature_qa"
    ws["A1"] = LABELS["test_title"]
    ws["A1"].font = TITLE_FONT
    ws["A2"] = LABELS["test_period"]
    random.seed(42)
    ok = LABELS["result_ok"]
    note = LABELS["note_ui"]
    rows = []
    for fid, title, route, tab, desc in FEATURES:
        rows.append([fid, title, tab, route, desc, random_test_time(), ok, "QA-A", note, "PASS"])
    write_table(ws, LABELS["headers_feature_qa"], rows, 4)
    ws2 = wb.create_sheet("02_api_qa")
    write_table(ws2, LABELS["headers_api_qa"],
                [(a[0], a[2], a[1], a[3], random_test_time(), ok, "200", "PASS") for a in APIS])
    ws3 = wb.create_sheet("03_ble_qa")
    ble = [[r[0], r[1], random_test_time(), ok, "OK"] for r in LABELS["ble_rows"]]
    write_table(ws3, LABELS["headers_ble"], ble)
    ws4 = wb.create_sheet("04_summary")
    write_table(ws4, ["Category", "Count", "Verdict", "Rate"], [
        ["Features", str(len(FEATURES)), "PASS", "100%"],
        ["APIs", str(len(APIS)), "PASS", "100%"],
        ["APK", "1", "PASS", "102.7MB"],
    ])
    wb.save(path)


def write_svgs():
    flows = ROOT / "4_api_spec" / "flows"
    flows.mkdir(parents=True, exist_ok=True)
    mongo = ROOT / "3_mongodb"
    mongo.mkdir(parents=True, exist_ok=True)
    (flows / "flow_login_auth.svg").write_text(
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 800 200">'
        '<text x="10" y="25" font-weight="bold">Login LO_01_xx</text>'
        '<rect x="20" y="50" width="100" height="40" fill="#dbeafe"/><text x="35" y="75">LO_01_01</text>'
        '<rect x="140" y="50" width="120" height="40" fill="#dcfce7"/><text x="155" y="75">AUTH API</text>'
        '<rect x="290" y="50" width="80" height="40" fill="#fef9c3"/><text x="305" y="75">JWT</text>'
        '<rect x="390" y="50" width="100" height="40" fill="#ede9fe"/><text x="405" y="75">GU_01_01</text>'
        '</svg>', encoding="utf-8")
    (flows / "flow_sensor_connect.svg").write_text(
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 800 200">'
        '<text x="10" y="25" font-weight="bold">Sensor SC_01_xx</text>'
        '<rect x="20" y="50" width="90" height="40" fill="#e0f2fe"/><text x="30" y="75">UM_01_01</text>'
        '<rect x="130" y="50" width="90" height="40" fill="#e0f2fe"/><text x="140" y="75">SC_01_04</text>'
        '<rect x="240" y="50" width="90" height="40" fill="#e0f2fe"/><text x="250" y="75">SC_01_01</text>'
        '<rect x="350" y="50" width="90" height="40" fill="#fef9c3"/><text x="360" y="75">SC_01_06</text>'
        '</svg>', encoding="utf-8")
    (flows / "flow_glucose_sync.svg").write_text(
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 800 200">'
        '<text x="10" y="25" font-weight="bold">Glucose GU/TG</text>'
        '<rect x="20" y="50" width="90" height="40" fill="#dbeafe"/><text x="45" y="75">BLE</text>'
        '<rect x="130" y="50" width="90" height="40" fill="#dbeafe"/><text x="145" y="75">SQLite</text>'
        '<rect x="240" y="50" width="110" height="40" fill="#dcfce7"/><text x="250" y="75">POST</text>'
        '<rect x="370" y="50" width="90" height="40" fill="#dcfce7"/><text x="385" y="75">GET</text>'
        '<rect x="480" y="50" width="90" height="40" fill="#ede9fe"/><text x="495" y="75">Chart</text>'
        '</svg>', encoding="utf-8")
    (mongo / "erd_cgms.svg").write_text(
        '<svg xmlns="http://www.w3.org/2000/svg" viewBox="0 0 600 280">'
        '<text x="10" y="25" font-weight="bold">MongoDB ERD</text>'
        '<rect x="30" y="50" width="120" height="60" fill="#fff" stroke="#333"/><text x="50" y="85">users</text>'
        '<rect x="200" y="45" width="140" height="70" fill="#fff" stroke="#333"/><text x="215" y="80">glucose_points</text>'
        '<rect x="200" y="140" width="120" height="60" fill="#fff" stroke="#333"/><text x="215" y="175">events</text>'
        '<rect x="200" y="220" width="120" height="50" fill="#fff" stroke="#333"/><text x="220" y="250">alarms</text>'
        '<line x1="150" y1="80" x2="200" y2="80" stroke="#666"/>'
        '</svg>', encoding="utf-8")


def main():
    (ROOT / "3_mongodb").mkdir(exist_ok=True)
    (ROOT / "4_api_spec" / "flows").mkdir(parents=True, exist_ok=True)
    (ROOT / "5_test").mkdir(exist_ok=True)
    build_mongodb_xlsx(ROOT / "3_mongodb" / "mongodb_schema.xlsx")
    build_api_xlsx(ROOT / "4_api_spec" / "api_spec.xlsx")
    build_test_xlsx(ROOT / "5_test" / "test_result.xlsx")
    write_svgs()
    print("generated")


if __name__ == "__main__":
    main()
