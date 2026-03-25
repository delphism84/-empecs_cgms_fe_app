import os
import re
import sys
from dataclasses import dataclass
from typing import List, Tuple

from openpyxl import load_workbook


@dataclass(frozen=True)
class ReqRow:
    row: int
    no: str
    page_id: str
    screen: str
    func: str
    desc: str
    etc: str
    j_251230: str
    k_260213: str


def _norm(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    s = re.sub(r"\r\n?", "\n", s)
    return s


def _bullets(txt: str) -> List[str]:
    txt = (txt or "").strip()
    if not txt:
        return []
    out: List[str] = []
    for line in re.split(r"\n+", txt):
        line = line.strip()
        if not line:
            continue
        line = re.sub(r"^\s*(?:•|\u2022|-|\*|\d+\.|\d+\))\s*", "", line).strip()
        if line:
            out.append(line)
    return out


def _prev_status_for(page_id: str, prev_text: str) -> Tuple[str, str]:
    if not prev_text:
        return ("unknown", "이전 요약(md) 없음")

    mapping = {
        "SC_03_01": ("done", "req260306.md: Scan&Connect 제거/정리 수정 완료"),
        "SC_01_01": ("done", "req260306.md: Scan&Connect 집중 수정 완료"),
        "SC_02_01": ("done", "req260306.md: Usage 불일치 수정 완료"),
        "SC_07_01": ("done", "req260306.md: Data Share 저장/목적지 근거 수정 완료"),
        "AR_01_01": ("done", "req260306.md: Alerts 저장/반영 수정 완료"),
        "Setting": (
            "partial",
            "req260306.md: Setting 일부(언어/시간/알림) 수정 완료, 일부(단위/접근성/센서 등) 미수정 표기",
        ),
        "LO_02_06": ("unknown", "req260306.md: 개발자 섹션 언급(상세 수정 여부 불명)"),
    }
    if page_id in mapping:
        return mapping[page_id]
    return ("not_covered", "req260306.md에 직접 언급 없음(이번 재분석 범위)")


def main() -> int:
    sys.stdout.reconfigure(encoding="utf-8")

    root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    xlsx = os.path.join(
        root,
        "req",
        "reqFinal",
        "디앤에스CGMS_앱_기능구성도_동작_확인_26_02_13 (1).xlsx",
    )
    prev_md = os.path.join(root, "req260306.md")
    out_md = os.path.join(root, "req", "reqFinal", "req_reanalysis_260313.md")

    if not os.path.exists(xlsx):
        print(f"ERROR: xlsx not found: {xlsx}")
        return 2

    prev_text = ""
    if os.path.exists(prev_md):
        with open(prev_md, "r", encoding="utf-8") as f:
            prev_text = f.read()

    wb = load_workbook(xlsx, data_only=True)
    if "Req" not in wb.sheetnames:
        print(f"ERROR: sheet 'Req' not found. sheets={wb.sheetnames}")
        return 2
    ws = wb["Req"]

    header_row = 4  # verified in this sheet

    # columns (1-indexed)
    COL_NO = 2
    COL_PAGE = 3
    COL_SCREEN = 4
    COL_FUNC = 5
    COL_DESC = 6
    COL_ETC = 7
    COL_J = 10  # 25.12.30 엠펙스 검토 사항
    COL_K = 11  # 26.02.13 엠펙스 검토 사항

    items: List[ReqRow] = []
    for r in range(header_row + 1, ws.max_row + 1):
        page_id = _norm(ws.cell(r, COL_PAGE).value)
        if page_id in ("", "페이지 ID"):
            continue

        j = _norm(ws.cell(r, COL_J).value)
        k = _norm(ws.cell(r, COL_K).value)
        if not (j or k):
            continue

        items.append(
            ReqRow(
                row=r,
                no=_norm(ws.cell(r, COL_NO).value),
                page_id=page_id,
                screen=_norm(ws.cell(r, COL_SCREEN).value),
                func=_norm(ws.cell(r, COL_FUNC).value),
                desc=_norm(ws.cell(r, COL_DESC).value),
                etc=_norm(ws.cell(r, COL_ETC).value),
                j_251230=j,
                k_260213=k,
            )
        )

    covered = sum(
        1 for it in items if _prev_status_for(it.page_id, prev_text)[0] != "not_covered"
    )

    lines: List[str] = []
    lines.append("# 기능 요구사항 재분석 + 수정 반영 여부 점검 (J/K 고객확인사항 기준)")
    lines.append("")
    lines.append("- 기준 파일: `req/reqFinal/디앤에스CGMS_앱_기능구성도_동작_확인_26_02_13 (1).xlsx`")
    lines.append("- 대상 컬럼: **J(25.12.30 엠펙스 검토 사항), K(26.02.13 엠펙스 검토 사항)**")
    lines.append("- 참고(이전 AI 요약): `req260306.md`")
    lines.append("")

    lines.append("## 요약")
    lines.append("")
    lines.append(f"- 전체 이슈/확인사항 행 수(J/K 비어있지 않음): **{len(items)}**")
    lines.append(f"- 이전 요약(req260306.md)과 직접 매칭(페이지ID 기준): **{covered}**")
    lines.append("")

    lines.append("## 1) 기능 요구사항 재분석 (고객확인사항 → 요구사항/수용기준)")
    lines.append("")

    for it in items:
        title = f"{it.page_id} · {it.screen}".strip()
        lines.append(f"### {title}")
        lines.append("")
        if it.func:
            lines.append(f"- **주요 기능(원문)**: {it.func}")
        if it.desc:
            lines.append(f"- **상세 설명(원문)**: {it.desc}")
        if it.etc:
            lines.append(f"- **기타(원문)**: {it.etc}")
        if it.j_251230:
            lines.append("- **고객확인사항 J(25.12.30)**:")
            for b in _bullets(it.j_251230):
                lines.append(f"  - {b}")
        if it.k_260213:
            lines.append("- **고객확인사항 K(26.02.13)**:")
            for b in _bullets(it.k_260213):
                lines.append(f"  - {b}")

        src = it.k_260213 or it.j_251230
        reqs: List[str] = []
        for b in _bullets(src):
            if re.search(r"안됨|되지 않음|오류|깨짐|보이지|반전|Timeout|네트워크", b):
                reqs.append(b)

        if reqs:
            lines.append("- **요구사항(재정의)**:")
            for rtxt in reqs:
                lines.append(f"  - {rtxt}")
            lines.append("- **수용기준(최소)**:")
            lines.append("  - 관련 화면에서 설정/토글/입력이 **저장**되고 앱 재실행 후에도 **반영**된다.")
            lines.append(
                "  - 네트워크 오류 시 사용자에게 **원인/재시도/오프라인 처리**가 일관되게 제공된다(무한 로딩/빈 화면 금지)."
            )
            lines.append(
                "  - 텍스트 깨짐/RTL 반전 등 **UI 깨짐이 재현되지 않는다**(기기/해상도 2종 이상)."
            )
        lines.append("")

    lines.append("## 2) 수정 반영 여부 분석 (이전 AI 요약 대비)")
    lines.append("")
    lines.append(
        "아래는 `req260306.md`에 적힌 “수정 완료/미수정” 표기와 **페이지ID 기준으로 매칭**한 결과입니다. "
        "(페이지ID가 요약에 없으면 “요약 미포함”으로 표시)"
    )
    lines.append("")

    label_map = {
        "done": "수정 완료(요약 표기)",
        "partial": "부분 수정(요약 표기)",
        "unknown": "불명",
        "not_covered": "요약 미포함",
    }
    for it in items:
        st, note = _prev_status_for(it.page_id, prev_text)
        lines.append(f"- **{it.page_id}**: {label_map[st]} — {note}")
    lines.append("")

    lines.append("## 3) 추가 수정 필요 후보 (J/K 기준)")
    lines.append("")
    lines.append(
        "다음 항목은 고객확인사항에 명확히 “안됨/미구현/저장 안됨/깨짐/오류”가 포함되어 있어, "
        "**실제 앱 반영 여부를 재검증하고 필요 시 수정해야 하는 후보**입니다."
    )
    lines.append("")

    for it in items:
        st, _ = _prev_status_for(it.page_id, prev_text)
        text = f"{it.k_260213}\n{it.j_251230}".strip()
        if not re.search(r"안됨|되지 않음|미\s*구현|저장\s*안됨|깨짐|오류|반전|보이지", text):
            continue
        if st not in ("not_covered", "unknown", "partial"):
            continue
        lines.append(f"- **{it.page_id} ({it.screen})**")
        for b in _bullets(it.k_260213 or it.j_251230):
            if re.search(r"안됨|되지 않음|미\s*구현|저장\s*안됨|깨짐|오류|반전|보이지", b):
                lines.append(f"  - {b}")

    os.makedirs(os.path.dirname(out_md), exist_ok=True)
    with open(out_md, "w", encoding="utf-8") as f:
        f.write("\n".join(lines).rstrip() + "\n")

    print(f"wrote: {out_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

